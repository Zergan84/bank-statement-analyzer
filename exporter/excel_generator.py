import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.transaction import Transaction
from analyzer.statistics import Statistics

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
INCOME_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
EXPENSE_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


class ExcelGenerator:
    def __init__(self, transactions: list[Transaction], stats: Statistics, source_pdf: str = ""):
        self.transactions = transactions
        self.stats = stats
        self.source_pdf = source_pdf

    def generate(self, output_path: str) -> str:
        wb = Workbook()
        self._create_transactions_sheet(wb)
        self._create_summary_sheet(wb)
        self._create_categories_sheet(wb)
        self._create_unknown_sheet(wb)
        wb.save(output_path)
        logger.info("Excel saved: %s", output_path)
        return output_path

    def _style_header(self, ws, cols: int) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

    def _auto_width(self, ws, cols: int, max_width: int = 40) -> None:
        for col in range(1, cols + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            adjusted = min(max_len + 3, max_width)
            ws.column_dimensions[get_column_letter(col)].width = adjusted

    def _create_transactions_sheet(self, wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Transactions"
        headers = [
            "№", "Date", "Type", "Sender/Recipient", "Description",
            "Amount", "Currency", "Category", "Confidence", "Comment", "Reference",
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        for i, tx in enumerate(self.transactions, start=1):
            date_str = tx.date.strftime("%d.%m.%Y") if tx.date else ""
            row_data = [
                i,
                date_str,
                tx.type.value if tx.type else "",
                tx.sender or tx.recipient or "",
                tx.description,
                tx.amount,
                tx.currency,
                tx.category.value if tx.category else "",
                tx.confidence_score,
                tx.comment,
                tx.reference,
            ]
            ws.append(row_data)
            row_num = i + 1
            fill = INCOME_FILL if tx.amount > 0 else EXPENSE_FILL
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = BORDER
                if col == 6:
                    cell.number_format = "#,##0.00"
                    cell.fill = fill
                if col == 2:
                    cell.alignment = Alignment(horizontal="center")
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(self.transactions) + 1}"
        ws.freeze_panes = "A2"
        self._auto_width(ws, len(headers))
        ws.sheet_properties.tabColor = "1F4E79"

    def _create_summary_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Summary")
        ws.sheet_properties.tabColor = "00B050"

        title_font = Font(bold=True, size=14, color="1F4E79")
        ws.cell(row=1, column=1, value="SUMMARY").font = title_font
        ws.cell(row=2, column=1, value=f"Source: {self.source_pdf}").font = Font(
            italic=True, color="666666"
        )

        labels = [
            ("Total transactions:", self.stats.total_count, None),
            ("Total Income:", round(self.stats.total_income, 2), "positive"),
            ("Total Expenses:", round(self.stats.total_expenses, 2), "negative"),
            ("Balance:", round(self.stats.balance, 2), "positive" if self.stats.balance >= 0 else "negative"),
            ("Categories used:", len(self.stats.category_stats), None),
        ]

        bold_font = Font(bold=True, size=11)
        for i, (label, value, style) in enumerate(labels, start=4):
            ws.cell(row=i, column=1, value=label).font = bold_font
            cell = ws.cell(row=i, column=2, value=value)
            cell.font = Font(size=11)
            if style == "positive" and isinstance(value, (int, float)):
                cell.font = Font(color="008000", bold=True, size=11)
                cell.number_format = "#,##0.00"
            elif style == "negative" and isinstance(value, (int, float)):
                cell.font = Font(color="FF0000", bold=True, size=11)
                cell.number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18

    def _create_categories_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Categories")
        ws.sheet_properties.tabColor = "FFC000"
        headers = ["Category", "Count", "Total"]
        ws.append(headers)
        self._style_header(ws, 3)

        for cat_name, data in sorted(self.stats.category_stats.items()):
            ws.append([cat_name, data["count"], round(data["total"], 2)])
            row_num = ws.max_row
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.border = BORDER

        self._auto_width(ws, 3)

    def _create_unknown_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Unknown")
        ws.sheet_properties.tabColor = "FF0000"
        headers = ["№", "Date", "Description", "Amount", "Currency", "Category", "Confidence"]
        ws.append(headers)
        self._style_header(ws, len(headers))

        for i, tx in enumerate(self.stats.unknown_transactions, start=1):
            ws.append([
                i,
                tx.date.strftime("%d.%m.%Y") if tx.date else "",
                tx.description,
                tx.amount,
                tx.currency,
                tx.category.value if tx.category else "",
                tx.confidence_score,
            ])
            row_num = i + 1
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = BORDER

        ws.auto_filter.ref = f"A1:G{len(self.stats.unknown_transactions) + 1}"
        self._auto_width(ws, len(headers))
